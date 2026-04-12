#!/usr/bin/env python3
"""
Build taxonomy JSON files from Jobs and Skills Australia Excel data.

This script reads the workbook used in the project and outputs:
1) a skill tag library derived from recurring task action phrases, and
2) an occupation role library derived from unique occupation names.

The output is intentionally deterministic so the front-end can consume stable
JSON files without parsing Excel in-browser.
"""

from __future__ import annotations

import json
import re
from collections import Counter
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

INPUT_PATH = Path("data/Occupation profiles data - February 2026.xlsx")
SKILLS_OUTPUT_PATH = Path("data/au-skills-taxonomy.json")
ROLES_OUTPUT_PATH = Path("data/au-role-taxonomy.json")


def normalize_whitespace(value: str) -> str:
    """
    Collapse repeated whitespace and trim surrounding spaces.

    Args:
        value: Raw text that may contain inconsistent spacing/newlines.

    Returns:
        A cleaned string with single spaces and no surrounding whitespace.
    """
    return re.sub(r"\s+", " ", value).strip()


def title_from_action_word(action_word: str) -> str:
    """
    Convert a task-leading action word into a human-friendly skill label.

    A mapping table handles frequent verbs found in occupation task data.
    For unknown words, an empty label is returned so the generated taxonomy
    stays concise and avoids noisy one-word verb artifacts.

    Args:
        action_word: First word extracted from a task statement.

    Returns:
        A concise skill label suitable for chip/tag UI display, or an empty
        string when the action word is not in the supported mapping.
    """
    mapping = {
        "assess": "Assessment",
        "assesses": "Assessment",
        "assessing": "Assessment",
        "assist": "Support Delivery",
        "assists": "Support Delivery",
        "assisting": "Support Delivery",
        "analyse": "Analysis",
        "analyses": "Analysis",
        "analyze": "Analysis",
        "answers": "Client Communication",
        "arrange": "Coordination",
        "arranges": "Coordination",
        "arranging": "Coordination",
        "authorise": "Resource Allocation",
        "authorising": "Resource Allocation",
        "check": "Quality Checking",
        "checks": "Quality Checking",
        "clean": "Cleaning Operations",
        "cleans": "Cleaning Operations",
        "cleaning": "Cleaning Operations",
        "collect": "Data Collection",
        "collecting": "Data Collection",
        "collects": "Data Collection",
        "consult": "Stakeholder Consultation",
        "consulting": "Stakeholder Consultation",
        "consults": "Stakeholder Consultation",
        "conduct": "Task Execution",
        "conducting": "Task Execution",
        "conducts": "Task Execution",
        "control": "Operational Control",
        "controls": "Operational Control",
        "coordinate": "Coordination",
        "co-ordinates": "Coordination",
        "design": "Design",
        "designs": "Design",
        "determine": "Decision Making",
        "determines": "Decision Making",
        "determining": "Decision Making",
        "develop": "Development",
        "develops": "Development",
        "developing": "Development",
        "direct": "Leadership",
        "directs": "Leadership",
        "ensure": "Compliance Management",
        "ensures": "Compliance Management",
        "ensuring": "Compliance Management",
        "examine": "Inspection",
        "examines": "Inspection",
        "examining": "Inspection",
        "identify": "Issue Identification",
        "identifies": "Issue Identification",
        "install": "Installation",
        "installs": "Installation",
        "maintain": "Maintenance",
        "maintains": "Maintenance",
        "maintaining": "Maintenance",
        "manage": "People and Operations Management",
        "manages": "People and Operations Management",
        "monitor": "Monitoring and Evaluation",
        "monitors": "Monitoring and Evaluation",
        "monitoring": "Monitoring and Evaluation",
        "operate": "Equipment and Systems Operation",
        "operates": "Equipment and Systems Operation",
        "operating": "Equipment and Systems Operation",
        "organise": "Work Organisation",
        "organises": "Work Organisation",
        "perform": "Operational Delivery",
        "performs": "Operational Delivery",
        "performing": "Operational Delivery",
        "plan": "Planning",
        "plans": "Planning",
        "planning": "Planning",
        "prepare": "Documentation and Reporting",
        "prepares": "Documentation and Reporting",
        "preparing": "Documentation and Reporting",
        "provide": "Service Delivery",
        "provides": "Service Delivery",
        "providing": "Service Delivery",
        "promote": "Promotion and Outreach",
        "promotes": "Promotion and Outreach",
        "receive": "Service Intake",
        "receives": "Service Intake",
        "record": "Record Keeping",
        "records": "Record Keeping",
        "research": "Research",
        "researches": "Research",
        "select": "Selection and Recruitment",
        "selecting": "Selection and Recruitment",
        "selects": "Selection and Recruitment",
        "set": "Target Setting",
        "sets": "Target Setting",
        "study": "Research",
        "studies": "Research",
        "supervise": "Supervision",
        "supervises": "Supervision",
        "undertake": "Operational Delivery",
        "undertakes": "Operational Delivery",
        "use": "Tool and Technology Use",
        "uses": "Tool and Technology Use",
    }
    normalized = action_word.lower().strip()
    if normalized in {"", "may"}:
        return ""
    return mapping.get(normalized, "")


def extract_occupations(workbook) -> list[str]:
    """
    Extract distinct occupation names from Table_1.

    Args:
        workbook: Loaded openpyxl workbook object.

    Returns:
        Sorted unique occupation names cleaned for JSON output.
    """
    sheet = workbook["Table_1"]
    occupations_to_employment: dict[str, int] = {}
    for row in sheet.iter_rows(min_row=8, values_only=True):
        if not row or len(row) < 2 or row[1] is None:
            continue
        occupation = normalize_whitespace(str(row[1]))
        if occupation:
            employed_raw = ""
            if len(row) > 2 and row[2] is not None:
                employed_raw = normalize_whitespace(str(row[2]))
            employed_digits = re.sub(r"[^\d]", "", employed_raw)
            employed_count = int(employed_digits) if employed_digits else 0
            previous_value = occupations_to_employment.get(occupation, 0)
            occupations_to_employment[occupation] = max(previous_value, employed_count)

    ranked = sorted(
        occupations_to_employment.items(),
        key=lambda item: (-item[1], item[0]),
    )
    return [occupation for occupation, _ in ranked]


def extract_task_action_words(workbook) -> Counter:
    """
    Count leading action words from occupation task statements in Table_3.

    Args:
        workbook: Loaded openpyxl workbook object.

    Returns:
        Counter keyed by task-leading action word frequency.
    """
    sheet = workbook["Table_3"]
    action_counter: Counter = Counter()
    for row in sheet.iter_rows(min_row=8, values_only=True):
        if not row or len(row) < 3 or row[2] is None:
            continue
        task = normalize_whitespace(str(row[2]))
        if not task:
            continue
        first_word = re.split(r"[^A-Za-z-]+", task)[0].lower()
        if first_word:
            action_counter[first_word] += 1
    return action_counter


def build_skill_labels(action_counter: Counter, minimum_frequency: int = 18) -> list[str]:
    """
    Convert action-word counts into a stable, de-duplicated skill label list.

    Args:
        action_counter: Frequency distribution of leading task action words.
        minimum_frequency: Minimum count threshold for inclusion.

    Returns:
        Sorted skill labels enriched with common transferability tags.
    """
    base_labels = {
        "Communication",
        "Team Collaboration",
        "Problem Solving",
        "Time Management",
        "Adaptability",
        "Attention to Detail",
    }

    for action_word, frequency in action_counter.items():
        if frequency < minimum_frequency:
            continue
        label = title_from_action_word(action_word)
        if label:
            base_labels.add(label)

    return sorted(base_labels)


def write_json(path: Path, values: Iterable[str]) -> None:
    """
    Persist a string list to JSON with UTF-8 encoding and stable indentation.

    Args:
        path: Output file location.
        values: Sequence of string entries to serialize.

    Returns:
        None.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(list(values), indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )


def main() -> None:
    """
    Execute end-to-end taxonomy generation from the source Excel workbook.

    Returns:
        None.
    """
    if not INPUT_PATH.exists():
        raise FileNotFoundError(f"Input workbook not found: {INPUT_PATH}")

    workbook = load_workbook(INPUT_PATH, read_only=True, data_only=True)
    occupations = extract_occupations(workbook)
    action_counter = extract_task_action_words(workbook)
    skill_labels = build_skill_labels(action_counter)

    write_json(SKILLS_OUTPUT_PATH, skill_labels)
    write_json(ROLES_OUTPUT_PATH, occupations)

    print(f"Generated {len(skill_labels)} skill tags -> {SKILLS_OUTPUT_PATH}")
    print(f"Generated {len(occupations)} role tags -> {ROLES_OUTPUT_PATH}")


if __name__ == "__main__":
    main()
